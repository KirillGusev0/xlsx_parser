# parser/services/importer.py
import logging
from openpyxl import load_workbook
from django.db import IntegrityError
from django.core.validators import validate_email
from django.core.exceptions import ValidationError
import time
from parser.models import EmailTask

logger = logging.getLogger(__name__)


class ImportResult:
    def __init__(self):
        self.total = 0
        self.created = 0
        self.skipped = 0
        self.errors = 0
        self.start_time = None
        self.end_time = None

    @property
    def duration(self):
        if self.start_time and self.end_time:
            return self.end_time - self.start_time
        return None


def validate_row(data: dict):
    if not data["external_id"]:
        raise ValueError("external_id is required")

    if not data["user_id"]:
        raise ValueError("user_id is required")

    try:
        int(data["user_id"])
    except Exception:
        raise ValueError("user_id must be integer")

    if not data["email"]:
        raise ValueError("email is required")

    try:
        validate_email(data["email"])
    except ValidationError:
        raise ValueError("invalid email")

    if not data["subject"]:
        raise ValueError("subject is required")

    if not data["message"]:
        raise ValueError("message is required")
    
def import_from_xlsx(file_path: str) -> ImportResult:
    
    result = ImportResult()
    
    result.start_time = time.time()

    workbook = load_workbook(filename=file_path, read_only=True)
    sheet = workbook.active

    rows = sheet.iter_rows(values_only=True)
    headers = next(rows)

    header_map = {name: idx for idx, name in enumerate(headers)}

    required_fields = ["external_id", "user_id", "email", "subject", "message"]

    for field in required_fields:
        if field not in header_map:
            raise ValueError(f"Missing required column: {field}")

    existing_ids = set(
        EmailTask.objects.values_list("external_id", flat=True)
    )

    for row in rows:
        result.total += 1

        try:
            data = {
                "external_id": row[header_map["external_id"]],
                "user_id": row[header_map["user_id"]],
                "email": row[header_map["email"]],
                "subject": row[header_map["subject"]],
                "message": row[header_map["message"]],
            }

            validate_row(data)

            if data["external_id"] in existing_ids:
                result.skipped += 1
                continue

            EmailTask.objects.create(**data)
            existing_ids.add(data["external_id"])

            result.created += 1
        except IntegrityError:
            result.skipped += 1

        except ValueError as e:
            logger.warning(f"Validation error: {row} - {e}")
            result.errors += 1

        except Exception as e:
            logger.error(f"Unexpected error: {row} - {e}")
            result.errors += 1

    result.end_time = time.time()
    
    return result